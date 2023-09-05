import os
import os.path
import shutil
from os import path, system

import gradio as gr
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

system("git pull")


class Product:
    def __init__(self, name: str, desciption: str, price: float):
        self.id = -1
        self.name = name
        self.description = desciption
        self.price = price


class Bundle:
    def __init__(self, name: str, description: str, products=None):
        if products is None:
            products: list[Product] = []
        self.name = name
        self.description = description
        self.products = products
        self.sum = -1

    def add_product(self, product: Product):
        self.products.append(product)

    def set_sum(self, sum: float):
        self.sum = sum

    def calculate_sum(self):
        for p in self.products:
            self.sum += p.price
        self.sum += 1


def increment_id(last_id):
    return last_id[:-4] + "0" * (4 - len(t := str(int(last_id.split("-")[2]) + 1))) + t


def send(price_list, shop_list, dataframe):
    return send_mode_init(price_list, shop_list), gr.update(
        value=load_dataframe(dataframe)
    )


def load_dataframe(dataframe):

    dataframe_data = []

    data = get_all_products()
    products = data[0]
    bundles = data[1]

    for product in products:
        dataframe_data.append([product.name, product.description, product.price])

    return dataframe_data


def on_price_list_change(price_list):
    if price_list:
        print(f"Price list change detected.\nCopying new price list to {os.getcwd()}")
        shutil.copy(price_list.name, "price_list.xlsx")


def get_all_products():

    if not os.path.exists("price_list.xlsx"):
        raise gr.Error("Keine Preisliste gefunden. Bitte hochladen!")

    price_wb: Workbook = load_workbook("price_list.xlsx")

    try:
        price_ws: Worksheet = price_wb["40495396"]
    except KeyError:
        raise gr.Error("Preisliste enthält falsche Daten")

    price_data = price_ws["A:E"]
    names = price_data[0]
    descriptions = price_data[4]
    prices = price_data[3]
    bundle = None

    bundles: list[Bundle] = []
    products: list[Product] = []

    for i in range(1, len(names) - 5):
        name = names[i].value
        price = prices[i].value
        description = descriptions[i].value

        if name is None:
            continue

        name = name.strip()
        if name == "Summe":
            bundle.calculate_sum()
            bundles.append(bundle)
            bundle = None
            continue
        if price is None:
            bundle: Bundle = Bundle(name, description)
            continue

        product = Product(name, description, price)

        if bundle:
            bundle.add_product(product)
        else:
            products.append(product)

    # Alle Einzelprodukte zu einer Liste zusammenfügen
    for bundle in bundles:
        products += bundle.products

    # Alle Duplikate entfernen
    unique_products = []
    unique_names = set()

    for product in products:
        if product.name not in unique_names:
            unique_products.append(product)
            unique_names.add(product.name)

    products = unique_products

    return products, bundles


def edit_and_save(price_list):
    print("SAVING DATAFRAME")


def send_mode_init(price_list, shop_list):

    shop_wb: Workbook = load_workbook("shop_template.xlsx")

    try:
        shop_ws: Worksheet = shop_wb["Artikel"]
    except KeyError:
        raise gr.Error("Preisliste enthält falsche Daten")

    new_id = increment_id("HW-CIT-0000")
    total_rows = 4

    data = get_all_products()
    products = data[0]
    bundles = data[1]

    print(products)

    for p in products:
        print(p.name)

    # IDS für die Einzelprodukte setzen
    for product in products:
        product.id = new_id
        new_id = increment_id(new_id)

    # Einzelprodukte in die Excel hinzufügen
    for product in products:
        shop_ws.cell(row=total_rows, column=2).value = product.id
        shop_ws.cell(row=total_rows, column=4).value = product.name
        shop_ws.cell(row=total_rows, column=5).value = product.description
        shop_ws.cell(row=total_rows, column=8).value = "Circ IT"
        shop_ws.cell(row=total_rows, column=9).value = "20"
        shop_ws.cell(row=total_rows, column=12).value = "Stück"
        shop_ws.cell(row=total_rows, column=16).value = product.price
        shop_ws.cell(row=total_rows, column=20).value = "EUR"
        shop_ws.cell(row=total_rows, column=21).value = "19"
        shop_ws.cell(row=total_rows, column=21).value = str(product.id) + ".png"
        shop_ws.cell(row=total_rows, column=25).value = "png"

        shop_ws.cell(row=total_rows, column=16).number_format = "#,##0.00€"
        shop_ws.cell(row=total_rows, column=21).number_format = "00"

        total_rows += 1

    # Alle Bundles durchgehen
    for bundle in bundles:

        ids = ""

        # IDs der Einzelprodukte bekommen
        for bundle_product in bundle.products:
            for product in products:
                if product.name == bundle_product.name:
                    ids += "\n" + str(product.id)

        shop_ws.cell(row=total_rows, column=1).value = "SET"
        shop_ws.cell(row=total_rows, column=2).value = new_id
        shop_ws.cell(row=total_rows, column=4).value = bundle.name
        shop_ws.cell(row=total_rows, column=5).value = bundle.description
        shop_ws.cell(row=total_rows, column=6).value = "enthält Artikel:" + ids
        shop_ws.cell(row=total_rows, column=8).value = "Circ IT"
        shop_ws.cell(row=total_rows, column=9).value = "20"
        shop_ws.cell(row=total_rows, column=12).value = "Stück"
        shop_ws.cell(row=total_rows, column=16).value = bundle.sum
        shop_ws.cell(row=total_rows, column=20).value = "EUR"
        shop_ws.cell(row=total_rows, column=21).value = "19"
        shop_ws.cell(row=total_rows, column=21).value = str(new_id) + ".png"
        shop_ws.cell(row=total_rows, column=25).value = "png"

        shop_ws.cell(row=total_rows, column=16).number_format = "#,##0.00€"
        shop_ws.cell(row=total_rows, column=21).number_format = "00"

        total_rows += 1

        new_id = increment_id(new_id)

    shop_wb.save("shop_list.xlsx")

    return gr.update(value=path.abspath("shop_list.xlsx"), visible=True)


modes_value = ["Zusammenführen (Nicht implementiert)", "Initiation"]
default_mode = modes_value[1]

settings_value = [
    "Neue Daten einpflegen",
    "Datenänderungen einpflegen",
    "Fehlerhafte Daten löschen",
]
default_settings = [settings_value[0], settings_value[1]]

with gr.Blocks() as app:
    with gr.Tab(label="Automatisierung Shopliste"):

        shop_list = gr.File(label="Shopliste", visible=False)
        price_list = gr.File(label="Neuste Preisliste", value="price_list.xlsx")
        send_button = gr.Button("Übertragen")

        output = gr.File(label="Neue Shopliste", visible=False)
    with gr.Tab(label="Manuelles ändern"):
        dataframe = gr.Dataframe(
            headers=["Name", "Beschreibung", "Preis (in €)"],
            datatype=["str", "str", "number"],
            label="Shopliste",
            interactive=True,
            col_count=(3, "fixed"),
            wrap=True,
        )
        dataframe.value = load_dataframe(dataframe)

    send_button.click(
        send,
        inputs=[price_list, shop_list, dataframe],
        outputs=[output, dataframe],
    )

    price_list.change(on_price_list_change, inputs=price_list)
    dataframe.input(edit_and_save, inputs=price_list)

app.queue()
app.launch(show_error=True, server_port=80, ssl_verify=True)
